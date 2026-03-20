# excel_export.py
# Generates formatted Excel workbooks from CMM inspection data.
# Creates a multi-sheet workbook with pass/fail summary,
# raw feature data, capability results, and trend data.
#
# New library: openpyxl
# openpyxl lets you create and format Excel files from Python.
# You build workbooks, add sheets, write cells, and apply styles.

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os


# ── Color Palette (Excel uses ARGB hex) ──────────────────────────
TEAL    = "FF00C896"
RED     = "FFFF4C4C"
AMBER   = "FFFFA500"
BLUE    = "FF4A90D9"
DARK    = "FF0F1117"
SURFACE = "FF1A1D27"
HEADER  = "FF2C3E50"
WHITE   = "FFFFFFFF"
LIGHT   = "FFF8FAFB"
GREEN_LIGHT  = "FFE8F8F2"
RED_LIGHT    = "FFFCE8E8"
AMBER_LIGHT  = "FFFFF3E0"


# ── Style Helpers ─────────────────────────────────────────────────

def header_font(size=11):
    return Font(name="Calibri", size=size, bold=True, color="FFFFFFFF")

def body_font(size=10, bold=False, color="FF2C3E50"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def header_fill(color=HEADER):
    return PatternFill(fill_type="solid", fgColor=color)

def cell_fill(color):
    return PatternFill(fill_type="solid", fgColor=color)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")

def thin_border():
    side = Side(style="thin", color="FFDDDDDD")
    return Border(left=side, right=side, top=side, bottom=side)

def write_header_row(ws, row: int, headers: list, col_widths: list = None):
    """Writes a styled header row."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = header_font()
        cell.fill      = header_fill()
        cell.alignment = center()
        cell.border    = thin_border()

    if col_widths:
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

def status_fill(status: str) -> PatternFill:
    """Returns a fill color based on pass/fail/warning status."""
    if status == "FAIL":    return cell_fill(RED_LIGHT)
    if status == "WARNING": return cell_fill(AMBER_LIGHT)
    return cell_fill(GREEN_LIGHT)

def rating_fill(rating: str) -> PatternFill:
    if rating == "Not Capable": return cell_fill(RED_LIGHT)
    if rating == "Marginal":    return cell_fill(AMBER_LIGHT)
    if rating == "Capable":     return cell_fill("FFE8F0FC")
    return cell_fill(GREEN_LIGHT)


# ── Sheet 1: Pass/Fail Summary ────────────────────────────────────

def write_summary_sheet(ws, results: list, source_file: str):
    """Writes the inspection summary sheet."""
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "CMM Inspection Summary"
    title_cell.font      = Font(name="Calibri", size=16, bold=True, color=HEADER[2:])
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 30

    # Source file
    ws.merge_cells("A2:F2")
    src_cell = ws["A2"]
    src_cell.value     = f"Source: {os.path.basename(source_file)}"
    src_cell.font      = body_font(size=10, color="FF7B8099")
    src_cell.alignment = left()
    ws.row_dimensions[2].height = 18

    # Stats row
    total    = len(results)
    passes   = sum(1 for r in results if r.is_pass)
    fails    = total - passes
    warnings = sum(1 for r in results if r.severity == "WARNING")
    rate     = round(passes / total * 100) if total > 0 else 0

    stats = [
        ("Total Features", total,      BLUE[2:]),
        ("Passed",         passes,     TEAL[2:]),
        ("Failed",         fails,      RED[2:]),
        ("Warnings",       warnings,   AMBER[2:]),
        ("Pass Rate",      f"{rate}%", TEAL[2:] if rate >= 80 else AMBER[2:]),
    ]

    for col, (label, value, color) in enumerate(stats, 1):
        label_cell = ws.cell(row=4, column=col, value=label)
        label_cell.font      = body_font(size=9, color="FF7B8099")
        label_cell.alignment = center()

        value_cell = ws.cell(row=5, column=col, value=value)
        value_cell.font      = Font(name="Calibri", size=18, bold=True, color=color)
        value_cell.alignment = center()
        ws.row_dimensions[5].height = 32

    # Feature detail table
    headers = ["Feature", "Type", "Condition", "Deviation",
               "Stated Tol", "Bonus", "Effective Tol", "Tol Used %", "Status", "Severity"]
    col_widths = [14, 18, 12, 12, 12, 10, 14, 12, 10, 12]
    write_header_row(ws, 7, headers, col_widths)

    for row_idx, r in enumerate(results, 8):
        values = [
            r.feature_name,
            r.feature_type,
            r.material_condition,
            round(r.deviation, 4),
            round(r.upper_tolerance, 4),
            round(r.bonus_tolerance, 4) if r.has_bonus else "-",
            round(r.effective_upper, 4),
            round(r.percent_used, 1),
            r.status,
            r.severity,
        ]
        fill = status_fill(r.severity)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font()
            cell.fill      = fill if col_idx >= 9 else (
                cell_fill(LIGHT) if row_idx % 2 == 0 else cell_fill(WHITE)
            )
            cell.alignment = center()
            cell.border    = thin_border()

        ws.row_dimensions[row_idx].height = 18


# ── Sheet 2: Raw Feature Data ─────────────────────────────────────

def write_raw_data_sheet(ws, results: list):
    """Writes the raw feature data sheet."""
    ws.title = "Raw Data"
    ws.sheet_view.showGridLines = False

    headers = ["Feature", "Type", "Nominal", "Actual", "Deviation",
               "Upper Tol", "Lower Tol", "Tol Used %", "Status"]
    col_widths = [14, 18, 10, 10, 12, 12, 12, 12, 10]
    write_header_row(ws, 1, headers, col_widths)

    for row_idx, r in enumerate(results, 2):
        values = [
            r.feature_name,
            r.feature_type,
            round(r.upper_tolerance + r.lower_tolerance, 4),
            round(r.deviation + r.upper_tolerance + r.lower_tolerance, 4),
            round(r.deviation, 4),
            round(r.upper_tolerance, 4),
            round(r.lower_tolerance, 4),
            round(r.percent_used, 1),
            r.status,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font()
            cell.fill      = cell_fill(LIGHT) if row_idx % 2 == 0 else cell_fill(WHITE)
            cell.alignment = center()
            cell.border    = thin_border()
        ws.row_dimensions[row_idx].height = 18


# ── Sheet 3: Capability ───────────────────────────────────────────

def write_capability_sheet(ws, cap_results: list):
    """Writes the Cp/Cpk capability summary sheet."""
    ws.title = "Capability"
    ws.sheet_view.showGridLines = False

    # Reference table
    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Industry Ratings"
    ws["A1"].font      = header_font(12)
    ws["A1"].fill      = header_fill()
    ws["A1"].alignment = center()

    ratings = [
        ("< 1.00",      "Not Capable", RED_LIGHT),
        ("1.00 - 1.33", "Marginal",    AMBER_LIGHT),
        ("1.33 - 1.67", "Capable",     "FFE8F0FC"),
        (">= 1.67",     "Excellent",   GREEN_LIGHT),
    ]
    for ri, (cpk_range, label, color) in enumerate(ratings, 2):
        for ci, val in enumerate([cpk_range, label], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = body_font(bold=(ci == 2))
            cell.fill      = cell_fill(color)
            cell.alignment = center()
            cell.border    = thin_border()
        ws.row_dimensions[ri].height = 18

    # Capability table
    headers = ["Feature", "Type", "Samples", "Mean Dev",
               "Std Dev", "Cp", "Cpk", "Rating"]
    col_widths = [14, 18, 10, 12, 12, 10, 10, 14]
    write_header_row(ws, 7, headers, col_widths)

    for row_idx, r in enumerate(cap_results, 8):
        values = [
            r.feature_name,
            r.feature_type,
            r.sample_count,
            round(r.mean, 4),
            round(r.std_dev, 4),
            r.cp,
            r.cpk,
            r.rating,
        ]
        fill = rating_fill(r.rating)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font()
            cell.fill      = fill if col_idx == 8 else (
                cell_fill(LIGHT) if row_idx % 2 == 0 else cell_fill(WHITE)
            )
            cell.alignment = center()
            cell.border    = thin_border()
        ws.row_dimensions[row_idx].height = 18

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet 4: Trend Data ───────────────────────────────────────────

def write_trend_sheet(ws, runs: list):
    """Writes the multi-run trend data sheet."""
    from trend import get_all_feature_names, get_feature_trend

    ws.title = "Trend Data"
    ws.sheet_view.showGridLines = False

    feature_names = get_all_feature_names(runs)
    run_names     = [r.run_name for r in runs]

    # Header row — run names
    ws.cell(row=1, column=1, value="Feature").font = header_font()
    ws.cell(row=1, column=1).fill      = header_fill()
    ws.cell(row=1, column=1).alignment = center()
    ws.column_dimensions["A"].width    = 16

    for col, name in enumerate(run_names, 2):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font      = header_font()
        cell.fill      = header_fill()
        cell.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Feature rows
    for row_idx, fname in enumerate(feature_names, 2):
        ws.cell(row=row_idx, column=1, value=fname).font = body_font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = left()

        trend = get_feature_trend(runs, fname)
        for col_idx, deviation in enumerate(trend, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=round(deviation, 4))
            cell.font      = body_font()
            cell.fill      = cell_fill(LIGHT) if row_idx % 2 == 0 else cell_fill(WHITE)
            cell.alignment = center()
            cell.border    = thin_border()
        ws.row_dimensions[row_idx].height = 18


# ── Main Export Function ──────────────────────────────────────────

def export_to_excel(
    output_path: str,
    results: list,
    source_file: str = "report",
    cap_results: list = None,
    runs: list = None,
) -> bool:
    """
    Creates a formatted Excel workbook with up to four sheets.

    Args:
        output_path  : Where to save the .xlsx file
        results      : List of EvaluationResult objects
        source_file  : Original report filename for display
        cap_results  : Optional Cp/Cpk results for capability sheet
        runs         : Optional RunData list for trend sheet

    Returns True if successful, False on error.
    """
    try:
        wb = openpyxl.Workbook()

        # Sheet 1 — Summary (replaces default empty sheet)
        ws_summary = wb.active
        write_summary_sheet(ws_summary, results, source_file)

        # Sheet 2 — Raw Data
        ws_raw = wb.create_sheet()
        write_raw_data_sheet(ws_raw, results)

        # Sheet 3 — Capability (only if cap_results provided)
        if cap_results:
            ws_cap = wb.create_sheet()
            write_capability_sheet(ws_cap, cap_results)

        # Sheet 4 — Trend Data (only if runs provided)
        if runs:
            ws_trend = wb.create_sheet()
            write_trend_sheet(ws_trend, runs)

        wb.save(output_path)
        return True

    except Exception as e:
        print(f"Excel export failed: {e}")
        return False